import csv
import os
from settings import SENDER_EMAIL, PASSWORD, DISPLAY_NAME
from smtplib import SMTP
import markdown
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from io import BytesIO
from database import *
import pandas as pd

async def get_msg(xlsxFile, template):
    content = await xlsxFile.read()
    df = pd.read_excel(BytesIO(content), header=0)
    data = df.to_dict(orient='records')
    headers = df.columns.tolist()
    # workbook = load_workbook(filename=BytesIO(content), data_only=True)
    # sheet = workbook.active
    # data = []
    # for row in sheet.iter_rows(values_only=True):
    #     data.append(row)
    # headers = list(data[0])
    # data.pop(0)
    print(data)
    print(headers)
    for row in data:
        body_string = template.body
        subject_string = template.subject
        for index, header in enumerate(headers):
            value = row[header]
            body_string = body_string.replace(f'{{{header}}}', value)
            subject_string = subject_string.replace(f'{{{header}}}', value)
        yield row['EMAIL'], body_string, subject_string


def confirm_attachments():
    file_contents = []
    file_names = []
    try:
        for filename in os.listdir('ATTACH'):

            entry = input(f"""TYPE IN 'Y' AND PRESS ENTER IF YOU CONFIRM T0 ATTACH {filename} 
                                    TO SKIP PRESS ENTER: """)
            confirmed = True if entry == 'Y' else False
            if confirmed:
                file_names.append(filename)
                with open(f'{os.getcwd()}/ATTACH/{filename}', "rb") as f:
                    content = f.read()
                file_contents.append(content)

        return {'names': file_names, 'contents': file_contents}
    except FileNotFoundError:
        print('No ATTACH directory found...')


async def send_emails(server: SMTP, template, xlsxFile):

    attachments = confirm_attachments()
    sent_count = 0

    async for receiver, message, subject in get_msg(xlsxFile, template):
        multipart_msg = MIMEMultipart("alternative")

        multipart_msg["Subject"] = subject
        multipart_msg["From"] = DISPLAY_NAME + f' <{SENDER_EMAIL}>'
        multipart_msg["To"] = receiver

        text = message
        html = markdown.markdown(text)

        part1 = MIMEText(text, "plain")
        part2 = MIMEText(html, "html")

        multipart_msg.attach(part1)
        multipart_msg.attach(part2)

        if attachments:
            for content, name in zip(attachments['contents'], attachments['names']):
                attach_part = MIMEBase('application', 'octet-stream')
                attach_part.set_payload(content)
                encoders.encode_base64(attach_part)
                attach_part.add_header('Content-Disposition',
                                       f"attachment; filename={name}")
                multipart_msg.attach(attach_part)

        try:
            server.sendmail(SENDER_EMAIL, receiver,
                            multipart_msg.as_string())
        except Exception as err:
            print(f'Problem occurend while sending to {receiver} ')
            print(err)
            input("PRESS ENTER TO CONTINUE")
        else:
            sent_count += 1

    print(f"Sent {sent_count} emails")


async def mail_controller(template, xlsxFile):
    host = "smtp.office365.com"
    port = 587  # TLS replaced SSL in 1999

    server = SMTP(host=host, port=port)
    server.connect(host=host, port=port)
    server.ehlo()
    server.starttls()
    server.ehlo()
    server.login(user=SENDER_EMAIL, password=PASSWORD)

    await send_emails(server, template, xlsxFile)

    server.quit()
    print("MAIL SEND")
    return {"message": "succeeded"}


# AAHNIK 2020
